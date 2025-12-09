"""
店铺详情页 API - 重构版本
使用服务层和工具类，大幅减少代码冗余
"""

import frappe
import json
from product_sales_planning.utils.response_utils import success_response, error_response
from product_sales_planning.utils.validation_utils import (
	parse_json_param,
	validate_required_params,
	validate_positive_integer,
	validate_doctype_exists,
	validate_month_format
)
from product_sales_planning.services.commodity_service import CommodityScheduleService


@frappe.whitelist()
def get_filter_options():
	"""获取筛选器选项"""
	try:
		tasks = frappe.get_all(
			"Schedule tasks",
			fields=["name"],
			order_by="creation desc"
		)

		stores = frappe.get_all(
			"Store List",
			fields=["name", "shop_name"],
			order_by="shop_name asc"
		)

		return success_response(data={
			"stores": stores,
			"tasks": tasks
		})

	except Exception as e:
		frappe.log_error(title="获取筛选器选项失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def get_store_commodity_data(store_id=None, task_id=None, brand=None, category=None,
                              start=0, page_length=20, search_term=None, view_mode="single"):
	"""
	查询商品规划数据

	Args:
		store_id: 店铺ID
		task_id: 任务ID
		brand: 品牌
		category: 类别
		start: 起始位置
		page_length: 每页条数
		search_term: 搜索关键词
		view_mode: 视图模式 ("single" | "multi")
	"""
	try:
		result = CommodityScheduleService.get_commodity_data(
			store_id=store_id,
			task_id=task_id,
			brand=brand,
			category=category,
			start=start,
			page_length=page_length,
			search_term=search_term,
			view_mode=view_mode
		)

		# 添加店铺和任务信息
		if store_id:
			store_info = frappe.get_value(
				"Store List",
				store_id,
				["name", "shop_name", "store_type", "region"],
				as_dict=True
			)
			result["store_info"] = store_info or {}
		else:
			result["store_info"] = {}

		if task_id:
			task_info = frappe.get_value(
				"Schedule tasks",
				task_id,
				["name", "task_name", "task_type", "start_date", "end_date", "status"],
				as_dict=True
			)
			result["task_info"] = task_info or {}
		else:
			result["task_info"] = {}

		# 获取审批状态和编辑权限
		if store_id and task_id:
			from product_sales_planning.planning_system.doctype.approval_workflow.approval_api import check_can_edit

			can_edit_result = check_can_edit(task_id, store_id)
			result["can_edit"] = can_edit_result.get("can_edit", True)
			result["edit_reason"] = can_edit_result.get("reason", "")

			# 获取审批状态
			tasks_store = frappe.db.get_value(
				"Tasks Store",
				{"parent": task_id, "store_name": store_id},
				["approval_status", "status"],
				as_dict=True
			)
			result["approval_status"] = tasks_store or {}
		else:
			result["can_edit"] = True
			result["edit_reason"] = ""
			result["approval_status"] = {}

		# 直接返回结果，不包装（前端期望直接获取数据）
		return result

	except Exception as e:
		frappe.log_error(title="查询商品规划数据失败", message=str(e))
		return {"error": str(e)}


@frappe.whitelist()
def bulk_insert_commodity_schedule(store_id, codes, task_id=None):
	"""批量插入商品计划"""
	try:
		# 解析参数
		codes = parse_json_param(codes, "商品代码列表")

		result = CommodityScheduleService.bulk_insert(
			store_id=store_id,
			task_id=task_id,
			codes=codes
		)
		return result

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="批量添加商品失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def batch_update_quantity(names, quantity):
	"""批量修改数量"""
	try:
		names = parse_json_param(names, "记录名称列表")
		result = CommodityScheduleService.batch_update_quantity(names, quantity)
		return result

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="批量修改失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def batch_delete_items(names):
	"""批量删除记录"""
	try:
		names = parse_json_param(names, "记录名称列表")
		result = CommodityScheduleService.batch_delete(names)
		return result

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="批量删除失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def batch_delete_by_codes(store_id, task_id, codes):
	"""根据产品编码批量删除记录（用于多月视图）"""
	try:
		codes = parse_json_param(codes, "商品代码列表")
		validate_required_params(
			{"store_id": store_id, "task_id": task_id},
			["store_id", "task_id"]
		)

		deleted_count = 0
		errors = []

		for code in codes:
			try:
				filters = {
					"code": code,
					"store_id": store_id,
					"task_id": task_id
				}

				records = frappe.get_all("Commodity Schedule", filters=filters, fields=["name"])

				for record in records:
					try:
						frappe.delete_doc("Commodity Schedule", record.name)
						deleted_count += 1
					except frappe.PermissionError:
						errors.append(f"产品 {code} 记录 {record.name}: 无权限删除")
					except Exception as inner_e:
						errors.append(f"产品 {code} 记录 {record.name}: {str(inner_e)}")

			except Exception as e:
				errors.append(f"产品 {code}: {str(e)}")
				frappe.log_error(f"删除产品记录失败: {code}", str(e))

		frappe.db.commit()

		msg = f"成功删除 {deleted_count} 条记录"
		if errors:
			msg += f"，{len(errors)} 条失败"

		return success_response(
			message=msg,
			count=deleted_count,
			errors=errors[:10]
		)

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="批量删除失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def update_line_item(name, field, value):
	"""更新单个字段"""
	try:
		frappe.db.set_value("Commodity Schedule", name, field, value)
		frappe.db.commit()
		return success_response(message="已保存")

	except Exception as e:
		frappe.log_error(title="更新字段失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def apply_mechanisms(store_id, mechanism_names, task_id=None):
	"""应用产品机制（批量添加机制中的所有产品）"""
	try:
		from product_sales_planning.utils.date_utils import get_next_n_months, get_month_first_day

		# 解析参数
		mechanism_names = parse_json_param(mechanism_names, "机制名称列表")
		validate_required_params({"store_id": store_id}, ["store_id"])
		validate_doctype_exists("Store List", store_id, "店铺")

		if task_id:
			validate_doctype_exists("Schedule tasks", task_id, "计划任务")

		inserted_count = 0
		skipped_count = 0
		errors = []

		# 获取下个月第一天
		next_month = get_next_n_months(n=1, include_current=False)[0]
		sub_date = get_month_first_day(next_month)

		for mech_name in mechanism_names:
			if not frappe.db.exists("Product Mechanism", mech_name):
				errors.append(f"机制 {mech_name} 不存在")
				continue

			try:
				mech_doc = frappe.get_doc("Product Mechanism", mech_name)

				if mech_doc.product_list:
					for item in mech_doc.product_list:
						product_code = item.name1
						default_qty = item.quantity or 1

						try:
							filters = {
								"store_id": store_id,
								"code": product_code,
								"sub_date": sub_date
							}
							if task_id:
								filters["task_id"] = task_id

							exists = frappe.db.exists("Commodity Schedule", filters)

							if not exists:
								new_doc = frappe.new_doc("Commodity Schedule")
								new_doc.store_id = store_id
								new_doc.code = product_code
								new_doc.task_id = task_id
								new_doc.quantity = default_qty
								new_doc.sub_date = sub_date
								new_doc.insert()
								inserted_count += 1
							else:
								skipped_count += 1

						except frappe.ValidationError as ve:
							errors.append(f"机制[{mech_name}]-产品[{product_code}]: {str(ve)}")
						except frappe.PermissionError:
							errors.append(f"机制[{mech_name}]-产品[{product_code}]: 无权限创建")
						except Exception as inner_e:
							errors.append(f"机制[{mech_name}]-产品[{product_code}]: {str(inner_e)}")

			except Exception as mech_e:
				errors.append(f"机制 {mech_name}: {str(mech_e)}")

		frappe.db.commit()

		msg = f"成功添加 {inserted_count} 条"
		if skipped_count > 0:
			msg += f"，跳过 {skipped_count} 条已存在记录"

		return success_response(
			message=msg,
			count=inserted_count,
			skipped=skipped_count,
			errors=errors[:10]
		)

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="应用机制失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def update_month_quantity(store_id, task_id, code, month, quantity):
	"""更新指定产品的某个月份的计划数量"""
	try:
		from product_sales_planning.utils.date_utils import get_month_first_day

		# 参数验证
		validate_required_params(
			{"store_id": store_id, "task_id": task_id, "code": code, "month": month},
			["store_id", "task_id", "code", "month"]
		)
		quantity = validate_positive_integer(quantity, "数量")
		validate_month_format(month)

		# 构造日期
		sub_date = get_month_first_day(month)

		# 查找记录
		filters = {
			"store_id": store_id,
			"task_id": task_id,
			"code": code,
			"sub_date": sub_date
		}

		existing = frappe.db.get_value("Commodity Schedule", filters, "name")

		if existing:
			doc = frappe.get_doc("Commodity Schedule", existing)
			doc.quantity = quantity
			doc.save()
		else:
			new_doc = frappe.new_doc("Commodity Schedule")
			new_doc.store_id = store_id
			new_doc.task_id = task_id
			new_doc.code = code
			new_doc.quantity = quantity
			new_doc.sub_date = sub_date
			new_doc.insert()

		frappe.db.commit()
		return success_response(message="已保存")

	except frappe.ValidationError as ve:
		frappe.db.rollback()
		return error_response(message=str(ve))
	except frappe.PermissionError:
		frappe.db.rollback()
		return error_response(message="无权限操作")
	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="更新月份数量失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def batch_update_quantities(store_id, task_id, updates):
	"""批量更新多个产品的多个月份的数量"""
	try:
		from product_sales_planning.utils.date_utils import get_month_first_day

		# 解析参数
		updates = parse_json_param(updates, "更新列表")
		validate_required_params(
			{"store_id": store_id, "task_id": task_id},
			["store_id", "task_id"]
		)

		updated_count = 0
		created_count = 0
		errors = []

		for update in updates:
			try:
				code = update.get("code")
				month = update.get("month")
				quantity = update.get("quantity")

				if not code or not month:
					errors.append(f"缺少必要参数: code={code}, month={month}")
					continue

				# 验证数量
				try:
					quantity = int(quantity)
					if quantity < 0:
						errors.append(f"产品 {code} 月份 {month}: 数量必须为正整数")
						continue
				except (ValueError, TypeError):
					errors.append(f"产品 {code} 月份 {month}: 数量格式错误")
					continue

				# 构造日期
				sub_date = get_month_first_day(month)

				# 查找记录
				filters = {
					"store_id": store_id,
					"task_id": task_id,
					"code": code,
					"sub_date": sub_date
				}

				existing = frappe.db.get_value("Commodity Schedule", filters, "name")

				if existing:
					frappe.db.set_value("Commodity Schedule", existing, "quantity", quantity)
					updated_count += 1
				else:
					new_doc = frappe.new_doc("Commodity Schedule")
					new_doc.store_id = store_id
					new_doc.task_id = task_id
					new_doc.code = code
					new_doc.quantity = quantity
					new_doc.sub_date = sub_date
					new_doc.insert()
					created_count += 1

			except Exception as inner_e:
				errors.append(f"产品 {code} 月份 {month}: {str(inner_e)}")

		frappe.db.commit()

		msg = f"成功更新 {updated_count} 条，创建 {created_count} 条"
		if errors:
			msg += f"，{len(errors)} 条失败"

		return success_response(
			message=msg,
			updated=updated_count,
			created=created_count,
			errors=errors[:10]
		)

	except Exception as e:
		frappe.db.rollback()
		frappe.log_error(title="批量更新数量失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def download_import_template(task_id=None):
	"""生成并下载Excel导入模板"""
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from frappe.utils.file_manager import save_file
		import io
		from datetime import datetime
		from product_sales_planning.utils.date_utils import get_next_n_months

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "商品计划导入模板"

		# 定义样式
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_font = Font(bold=True, color="FFFFFF", size=11)
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 生成月份列
		months = get_next_n_months(n=4, include_current=False)

		# 设置表头
		headers = ['产品编码', '产品名称'] + months
		ws.append(headers)

		# 设置表头样式
		for col_num in range(1, len(headers) + 1):
			cell = ws.cell(row=1, column=col_num)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.border = border

		# 添加示例数据
		example_row1 = ['PROD001', '示例商品A'] + [100 + i * 10 for i in range(len(months))]
		example_row2 = ['PROD002', '示例商品B'] + [50 + i * 5 for i in range(len(months))]

		for row_data in [example_row1, example_row2]:
			ws.append(row_data)
			for col_num in range(1, len(headers) + 1):
				cell = ws.cell(row=ws.max_row, column=col_num)
				cell.border = border
				if col_num > 2:
					cell.alignment = Alignment(horizontal='right', vertical='center')

		# 设置列宽
		ws.column_dimensions['A'].width = 15
		ws.column_dimensions['B'].width = 25
		for i in range(len(months)):
			col_letter = openpyxl.utils.get_column_letter(i + 3)
			ws.column_dimensions[col_letter].width = 12

		# 添加说明工作表
		ws_info = wb.create_sheet("填写说明")
		instructions = [
			["商品计划导入模板使用说明", ""],
			["", ""],
			["1. 基本要求", ""],
			["", "• 产品编码必须在系统中存在"],
			["", "• 产品名称仅用于参考，不影响导入"],
			["", "• 月份格式支持：2025-01、202501、2025/01"],
			["", "• 数量必须为整数，空值或0将被跳过"],
			["", ""],
			["2. 导入规则", ""],
			["", "• 如果记录已存在（相同店铺+任务+产品+月份），将更新数量"],
			["", "• 如果记录不存在，将创建新记录"],
			["", "• 导入前请确保已选择店铺和计划任务"],
			["", ""],
			["3. 注意事项", ""],
			["", "• 请勿修改表头行"],
			["", "• 建议单次导入不超过1000行数据"],
			["", "• 导入完成后请检查导入结果"],
		]

		for row_data in instructions:
			ws_info.append(row_data)

		ws_info.column_dimensions['A'].width = 20
		ws_info.column_dimensions['B'].width = 50
		title_cell = ws_info['A1']
		title_cell.font = Font(bold=True, size=14, color="4472C4")

		# 保存到内存
		file_content = io.BytesIO()
		wb.save(file_content)
		file_content.seek(0)

		filename = f"commodity_plan_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

		file_doc = save_file(
			fname=filename,
			content=file_content.read(),
			dt=None,
			dn=None,
			is_private=0
		)

		return success_response(
			file_url=file_doc.file_url,
			file_name=filename
		)

	except Exception as e:
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="生成导入模板失败", message=error_msg)
		return error_response(message=f"生成模板失败: {str(e)}")


@frappe.whitelist()
def import_commodity_data(store_id, task_id, file_url):
	"""从Excel导入商品计划数据"""
	try:
		import openpyxl
		from frappe.utils.file_manager import get_file_path
		from product_sales_planning.utils.date_utils import parse_month_string, get_month_first_day

		validate_required_params(
			{"store_id": store_id, "task_id": task_id},
			["store_id", "task_id"]
		)

		# 获取文件路径
		try:
			file_path = get_file_path(file_url)
		except Exception as e:
			return error_response(message=f"无法获取文件: {str(e)}")

		# 读取Excel文件
		try:
			wb = openpyxl.load_workbook(file_path, data_only=True)
			ws = wb.active
		except Exception as e:
			return error_response(message=f"无法读取Excel文件: {str(e)}")

		# 读取表头
		headers = []
		for cell in ws[1]:
			if cell.value:
				headers.append(str(cell.value).strip())
			else:
				headers.append("")

		if len(headers) < 3:
			return error_response(message="Excel格式错误：至少需要3列（产品编码、产品名称、月份数据）")

		# 解析月份列
		month_columns = [h for h in headers[2:] if h]

		if not month_columns:
			return error_response(message="Excel格式错误：未找到月份列")

		inserted_count = 0
		updated_count = 0
		errors = []
		skipped_count = 0

		# 从第2行开始读取数据
		for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			if not row or not row[0]:
				skipped_count += 1
				continue

			try:
				product_code = str(row[0]).strip()
			except:
				errors.append(f"第{row_idx}行: 产品编码格式错误")
				continue

			# 验证产品是否存在
			if not frappe.db.exists("Product List", product_code):
				errors.append(f"第{row_idx}行: 产品编码 {product_code} 不存在")
				continue

			# 处理每个月份的数据
			for col_idx, month_str in enumerate(month_columns):
				try:
					if len(row) <= 2 + col_idx:
						continue

					quantity_value = row[2 + col_idx]

					if quantity_value is None or quantity_value == '' or quantity_value == 0:
						continue

					try:
						quantity = int(float(quantity_value))
					except:
						errors.append(f"第{row_idx}行-{month_str}: 数量格式错误 ({quantity_value})")
						continue

					# 解析月份
					parsed_month = parse_month_string(month_str)
					if not parsed_month:
						errors.append(f"第{row_idx}行: 月份格式错误 ({month_str})")
						continue

					sub_date = get_month_first_day(parsed_month)

					# 检查记录是否存在
					filters = {
						"store_id": store_id,
						"task_id": task_id,
						"code": product_code,
						"sub_date": sub_date
					}

					existing = frappe.db.get_value("Commodity Schedule", filters, "name")

					if existing:
						frappe.db.set_value("Commodity Schedule", existing, "quantity", quantity)
						updated_count += 1
					else:
						new_doc = frappe.new_doc("Commodity Schedule")
						new_doc.store_id = store_id
						new_doc.task_id = task_id
						new_doc.code = product_code
						new_doc.quantity = quantity
						new_doc.sub_date = sub_date
						new_doc.insert()
						inserted_count += 1

				except Exception as inner_e:
					errors.append(f"第{row_idx}行-{month_str}: {str(inner_e)}")

		frappe.db.commit()

		msg = f"成功导入 {inserted_count} 条，更新 {updated_count} 条"
		if skipped_count > 0:
			msg += f"，跳过 {skipped_count} 行空数据"

		return success_response(
			message=msg,
			inserted=inserted_count,
			updated=updated_count,
			skipped=skipped_count,
			errors=errors[:20]
		)

	except Exception as e:
		frappe.db.rollback()
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="Excel导入失败", message=error_msg)
		return error_response(message=f"导入失败: {str(e)}")


# 以下方法保持不变，因为它们的逻辑较为复杂且特定
# batch_update_month_quantities, import_mechanism_excel, download_mechanism_template, get_approval_status

@frappe.whitelist()
def export_commodity_data(store_id=None, task_id=None):
	"""导出商品计划数据到Excel"""
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from frappe.utils.file_manager import save_file
		import io
		from datetime import datetime

		# 获取数据
		result = CommodityScheduleService.get_commodity_data(
			store_id=store_id,
			task_id=task_id,
			start=0,
			page_length=10000,  # 导出所有数据
			view_mode="multi"
		)

		if not result.get("data"):
			return error_response(message="没有数据可导出")

		data = result["data"]
		months = result.get("months", [])

		# 创建工作簿
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "商品计划数据"

		# 定义样式
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_font = Font(bold=True, color="FFFFFF", size=11)
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 设置表头
		headers = ['产品编码', '产品名称', '规格', '品牌', '类别'] + months
		ws.append(headers)

		# 设置表头样式
		for col_num in range(1, len(headers) + 1):
			cell = ws.cell(row=1, column=col_num)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.border = border

		# 写入数据
		for item in data:
			row_data = [
				item.get('code', ''),
				item.get('name1', ''),
				item.get('specifications', ''),
				item.get('brand', ''),
				item.get('category', '')
			]

			# 添加月份数据
			for month in months:
				month_data = item.get('months', {}).get(month, {})
				quantity = month_data.get('quantity', 0) if month_data else 0
				row_data.append(quantity)

			ws.append(row_data)

			# 设置数据行样式
			for col_num in range(1, len(headers) + 1):
				cell = ws.cell(row=ws.max_row, column=col_num)
				cell.border = border
				if col_num > 5:  # 月份列右对齐
					cell.alignment = Alignment(horizontal='right', vertical='center')

		# 设置列宽
		ws.column_dimensions['A'].width = 15  # 产品编码
		ws.column_dimensions['B'].width = 25  # 产品名称
		ws.column_dimensions['C'].width = 12  # 规格
		ws.column_dimensions['D'].width = 12  # 品牌
		ws.column_dimensions['E'].width = 12  # 类别
		for i in range(len(months)):
			col_letter = openpyxl.utils.get_column_letter(i + 6)
			ws.column_dimensions[col_letter].width = 12

		# 保存到内存
		file_content = io.BytesIO()
		wb.save(file_content)
		file_content.seek(0)

		# 生成文件名
		filename_parts = ["commodity_plan_export"]
		if store_id:
			filename_parts.append(f"store_{store_id}")
		if task_id:
			filename_parts.append(f"task_{task_id}")
		filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
		filename = "_".join(filename_parts) + ".xlsx"

		# 保存文件
		file_doc = save_file(
			fname=filename,
			content=file_content.read(),
			dt=None,
			dn=None,
			is_private=0
		)

		return success_response(
			file_url=file_doc.file_url,
			file_name=filename,
			record_count=len(data)
		)

	except Exception as e:
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="导出Excel失败", message=error_msg)
		return error_response(message=f"导出失败: {str(e)}")


@frappe.whitelist()
def get_tasks_store_status(task_id, store_id):
	"""获取任务店铺状态信息"""
	try:
		# 获取Tasks Store记录
		tasks_store = frappe.db.get_value("Tasks Store", {
			"parent": task_id,
			"store_name": store_id
		}, ["name", "status", "approval_status", "current_approval_step", "can_edit", 
			"rejection_reason", "submitted_by", "current_approver", "workflow_id"], as_dict=True)

		if not tasks_store:
			return error_response(message="未找到对应的任务店铺记录")

		# 检查当前用户是否是当前审批人
		current_user = frappe.session.user
		is_current_approver = tasks_store.get("current_approver") == current_user

		# 检查是否可以撤回（仅限提交人且在待审批状态）
		can_withdraw = (
			tasks_store.get("submitted_by") == current_user and 
			tasks_store.get("approval_status") == "待审批"
		)

		# 添加计算字段
		tasks_store["is_current_approver"] = 1 if is_current_approver else 0
		tasks_store["can_withdraw"] = 1 if can_withdraw else 0

		return success_response(data=tasks_store)

	except Exception as e:
		frappe.log_error(title="获取任务店铺状态失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def get_approval_status(task_id, store_id):
	"""获取审批状态和流程信息"""
	try:
		from product_sales_planning.planning_system.doctype.approval_workflow.approval_api import (
			get_workflow_for_task_store,
			get_approval_history,
			check_can_edit
		)

		workflow_info = get_workflow_for_task_store(task_id, store_id)
		history_info = get_approval_history(task_id, store_id)
		edit_info = check_can_edit(task_id, store_id)

		current_user = frappe.session.user
		user_roles = frappe.get_roles(current_user)

		can_approve = False
		if workflow_info.get("has_workflow") and workflow_info.get("current_state"):
			current_state = workflow_info["current_state"]
			if current_state.get("approval_status") == "待审批":
				current_step = current_state.get("current_step", 0)
				if current_step > 0 and workflow_info.get("workflow"):
					steps = workflow_info["workflow"].get("steps", [])
					if current_step <= len(steps):
						required_role = steps[current_step - 1].get("approver_role")
						can_approve = required_role in user_roles or "System Manager" in user_roles

		return success_response(data={
			"workflow": workflow_info,
			"history": history_info.get("data", []),
			"can_edit": edit_info.get("can_edit", True),
			"edit_reason": edit_info.get("reason", ""),
			"can_approve": can_approve,
			"user_roles": user_roles
		})

	except Exception as e:
		frappe.log_error(title="获取审批状态失败", message=str(e))
		return error_response(message=str(e))


@frappe.whitelist()
def get_product_list_for_dialog(store_id=None, task_id=None):
	"""获取产品列表用于添加商品对话框"""
	try:
		# 获取所有产品
		products = frappe.get_all(
			"Product List",
			fields=["name", "code", "name1", "brand", "category", "specifications"],
			order_by="code asc",
			limit_page_length=1000
		)

		# 获取已添加的商品编码（如果提供了 store_id 和 task_id）
		existing_codes = set()
		if store_id and task_id:
			existing = frappe.get_all(
				"Commodity Schedule",
				filters={"store_id": store_id, "task_id": task_id},
				fields=["code"],
				pluck="code"
			)
			existing_codes = set(existing)

		# 格式化产品数据，添加 commodity_code 和 commodity_name 字段以兼容前端
		formatted_products = []
		for product in products:
			formatted_products.append({
				"name": product.name,
				"code": product.code,
				"commodity_code": product.code,  # 兼容字段
				"commodity_name": product.name1,  # 兼容字段
				"name1": product.name1,
				"brand": product.brand or "",
				"category": product.category or "",
				"specifications": product.specifications or "",
				"is_added": product.code in existing_codes
			})

		return success_response(data=formatted_products)

	except Exception as e:
		frappe.log_error(title="获取产品列表失败", message=str(e))
		return error_response(message=str(e))
