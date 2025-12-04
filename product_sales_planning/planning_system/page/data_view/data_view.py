import frappe
from frappe.utils import getdate, today, format_datetime
import json


@frappe.whitelist()
def get_data_view(filters=None, page=1, page_size=50, sort_by=None, sort_order="asc"):
	"""
	获取跨店铺商品数据列表

	参数:
		filters: 筛选条件字典
			- task_ids: 任务ID列表
			- store_ids: 店铺ID列表
			- product_codes: 商品编码列表
			- start_date: 开始日期
			- end_date: 结束日期
			- channels: 渠道列表
			- approval_statuses: 审批状态列表
			- submission_statuses: 提交状态列表
		page: 页码（默认1）
		page_size: 每页数量（默认50）
		sort_by: 排序字段
		sort_order: 排序方向（asc/desc）

	返回:
		{
			"data": [...],  # 数据列表
			"total": 1000,  # 总记录数
			"page": 1,
			"page_size": 50,
			"stats": {
				"total_stores": 50,
				"total_products": 200,
				"total_quantity": 10000,
				"completed_stores": 30,
				"pending_stores": 15,
				"rejected_stores": 5
			}
		}
	"""
	try:
		# 解析过滤器参数
		if isinstance(filters, str):
			filters = json.loads(filters) if filters else {}
		elif filters is None:
			filters = {}

		# 构建 WHERE 条件
		conditions = ["1=1"]
		values = {}

		# 任务筛选
		if filters.get("task_ids"):
			task_ids = filters["task_ids"]
			if isinstance(task_ids, str):
				task_ids = [task_ids]
			elif not isinstance(task_ids, list):
				task_ids = [task_ids]
			conditions.append("st.name IN %(task_ids)s")
			values["task_ids"] = task_ids

		# 店铺筛选
		if filters.get("store_ids"):
			store_ids = filters["store_ids"]
			if isinstance(store_ids, str):
				store_ids = [store_ids]
			elif not isinstance(store_ids, list):
				store_ids = [store_ids]
			conditions.append("cs.store_id IN %(store_ids)s")
			values["store_ids"] = store_ids

		# 商品筛选
		if filters.get("product_codes"):
			product_codes = filters["product_codes"]
			if isinstance(product_codes, str):
				product_codes = [product_codes]
			elif not isinstance(product_codes, list):
				product_codes = [product_codes]
			conditions.append("cs.code IN %(product_codes)s")
			values["product_codes"] = product_codes

		# 货品计划日期筛选（sub_date）
		if filters.get("plan_date"):
			conditions.append("DATE(cs.sub_date) = %(plan_date)s")
			values["plan_date"] = filters["plan_date"]

		# 渠道筛选
		if filters.get("channels"):
			channels = filters["channels"]
			if isinstance(channels, str):
				channels = [channels]
			elif not isinstance(channels, list):
				channels = [channels]
			conditions.append("sl.channel IN %(channels)s")
			values["channels"] = channels

		# 审批状态筛选
		if filters.get("approval_statuses"):
			approval_statuses = filters["approval_statuses"]
			if isinstance(approval_statuses, str):
				approval_statuses = [approval_statuses]
			elif not isinstance(approval_statuses, list):
				approval_statuses = [approval_statuses]
			conditions.append("ts.approval_status IN %(approval_statuses)s")
			values["approval_statuses"] = approval_statuses

		# 提交状态筛选
		if filters.get("submission_statuses"):
			submission_statuses = filters["submission_statuses"]
			if isinstance(submission_statuses, str):
				submission_statuses = [submission_statuses]
			elif not isinstance(submission_statuses, list):
				submission_statuses = [submission_statuses]
			conditions.append("ts.status IN %(submission_statuses)s")
			values["submission_statuses"] = submission_statuses

		where_clause = " AND ".join(conditions)

		# 排序字段映射
		sort_field_map = {
			"shop_name": "sl.shop_name",
			"channel": "sl.channel",
			"code": "cs.code",
			"product_name": "pl.name1",
			"quantity": "cs.quantity",
			"sub_date": "cs.sub_date",
			"approval_status": "ts.approval_status",
			"user": "ts.user"
		}

		order_by_clause = ""
		if sort_by and sort_by in sort_field_map:
			order_by_clause = f"ORDER BY {sort_field_map[sort_by]} {sort_order.upper()}"
		else:
			order_by_clause = "ORDER BY cs.sub_date DESC"

		# 计算偏移量
		offset = (int(page) - 1) * int(page_size)

		# 查询数据
		query = f"""
			SELECT
				cs.name,
				cs.store_id,
				sl.shop_name,
				sl.channel,
				cs.code,
				pl.name1 as product_name,
				pl.specifications,
				pl.brand,
				pl.category,
				cs.quantity,
				cs.sub_date,
				ts.approval_status,
				ts.status as submission_status,
				ts.user,
				st.type as task_type,
				st.name as task_id,
				st.start_date,
				st.end_date
			FROM `tabCommodity Schedule` cs
			LEFT JOIN `tabStore List` sl ON cs.store_id = sl.name
			LEFT JOIN `tabProduct List` pl ON cs.code = pl.name
			LEFT JOIN `tabSchedule tasks` st ON cs.task_id = st.name
			LEFT JOIN `tabTasks Store` ts ON ts.parent = st.name AND ts.store_name = cs.store_id
			WHERE {where_clause}
			{order_by_clause}
			LIMIT %(limit)s OFFSET %(offset)s
		"""

		values["limit"] = int(page_size)
		values["offset"] = offset

		data = frappe.db.sql(query, values, as_dict=True)

		# 查询总记录数
		count_query = f"""
			SELECT COUNT(*) as total
			FROM `tabCommodity Schedule` cs
			LEFT JOIN `tabStore List` sl ON cs.store_id = sl.name
			LEFT JOIN `tabProduct List` pl ON cs.code = pl.name
			LEFT JOIN `tabSchedule tasks` st ON cs.task_id = st.name
			LEFT JOIN `tabTasks Store` ts ON ts.parent = st.name AND ts.store_name = cs.store_id
			WHERE {where_clause}
		"""

		total_result = frappe.db.sql(count_query, values, as_dict=True)
		total = total_result[0]["total"] if total_result else 0

		# 计算统计信息
		stats_query = f"""
			SELECT
				COUNT(DISTINCT cs.store_id) as total_stores,
				COUNT(DISTINCT cs.code) as total_products,
				SUM(cs.quantity) as total_quantity,
				COUNT(DISTINCT CASE WHEN ts.approval_status = '已通过' THEN cs.store_id END) as completed_stores,
				COUNT(DISTINCT CASE WHEN ts.approval_status = '待审批' THEN cs.store_id END) as pending_stores,
				COUNT(DISTINCT CASE WHEN ts.approval_status = '已驳回' THEN cs.store_id END) as rejected_stores
			FROM `tabCommodity Schedule` cs
			LEFT JOIN `tabStore List` sl ON cs.store_id = sl.name
			LEFT JOIN `tabProduct List` pl ON cs.code = pl.name
			LEFT JOIN `tabSchedule tasks` st ON cs.task_id = st.name
			LEFT JOIN `tabTasks Store` ts ON ts.parent = st.name AND ts.store_name = cs.store_id
			WHERE {where_clause}
		"""

		stats_result = frappe.db.sql(stats_query, values, as_dict=True)
		stats = stats_result[0] if stats_result else {
			"total_stores": 0,
			"total_products": 0,
			"total_quantity": 0,
			"completed_stores": 0,
			"pending_stores": 0,
			"rejected_stores": 0
		}

		return {
			"status": "success",
			"data": data,
			"total": total,
			"page": int(page),
			"page_size": int(page_size),
			"stats": stats
		}

	except Exception as e:
		frappe.log_error(title="获取数据查看失败", message=str(e))
		return {
			"status": "error",
			"message": str(e),
			"data": [],
			"total": 0,
			"page": 1,
			"page_size": 50,
			"stats": {
				"total_stores": 0,
				"total_products": 0,
				"total_quantity": 0,
				"completed_stores": 0,
				"pending_stores": 0,
				"rejected_stores": 0
			}
		}


@frappe.whitelist()
def get_data_view_filter_options():
	"""获取数据查看页面的筛选器选项"""
	try:
		# 获取所有开启中的任务
		tasks = frappe.get_all(
			"Schedule tasks",
			filters={"status": "开启中"},
			fields=["name", "type", "start_date", "end_date"],
			order_by="creation desc"
		)

		# 获取所有店铺
		stores = frappe.get_all(
			"Store List",
			fields=["name", "shop_name", "channel"],
			order_by="shop_name asc"
		)

		# 获取所有商品
		products = frappe.get_all(
			"Product List",
			fields=["name", "name1", "code", "brand", "category"],
			order_by="name1 asc",
			limit_page_length=1000  # 限制数量，避免数据过多
		)

		# 获取所有渠道
		channels = frappe.db.sql("""
			SELECT DISTINCT channel
			FROM `tabStore List`
			WHERE channel IS NOT NULL AND channel != ''
			ORDER BY channel
		""", as_dict=True)

		return {
			"status": "success",
			"tasks": tasks,
			"stores": stores,
			"products": products,
			"channels": [c["channel"] for c in channels],
			"approval_statuses": ["待审批", "已通过", "已驳回"],
			"submission_statuses": ["未开始", "已提交"]
		}

	except Exception as e:
		frappe.log_error(title="获取筛选选项失败", message=str(e))
		return {
			"status": "error",
			"message": str(e),
			"tasks": [],
			"stores": [],
			"products": [],
			"channels": [],
			"approval_statuses": [],
			"submission_statuses": []
		}


@frappe.whitelist()
def export_data_view(filters=None):
	"""导出数据为 Excel"""
	try:
		import openpyxl
		from openpyxl.styles import Font, Alignment, PatternFill
		from frappe.utils import get_site_path
		import os

		# 获取所有数据（不分页）
		result = get_data_view(filters=filters, page=1, page_size=999999)

		if result["status"] != "success":
			return {"status": "error", "message": "获取数据失败"}

		data = result["data"]

		# 创建工作簿
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "数据查看"

		# 写入表头
		headers = [
			"店铺名称", "渠道", "商品编码", "商品名称", "规格", "品牌", "类别",
			"数量", "提交时间", "审批状态", "提交状态", "负责人", "任务类型",
			"任务开始日期", "任务结束日期"
		]

		for col_idx, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_idx, value=header)
			cell.font = Font(bold=True)
			cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
			cell.alignment = Alignment(horizontal="center", vertical="center")

		# 写入数据
		for row_idx, row_data in enumerate(data, 2):
			ws.cell(row=row_idx, column=1, value=row_data.get("shop_name", ""))
			ws.cell(row=row_idx, column=2, value=row_data.get("channel", ""))
			ws.cell(row=row_idx, column=3, value=row_data.get("code", ""))
			ws.cell(row=row_idx, column=4, value=row_data.get("product_name", ""))
			ws.cell(row=row_idx, column=5, value=row_data.get("specifications", ""))
			ws.cell(row=row_idx, column=6, value=row_data.get("brand", ""))
			ws.cell(row=row_idx, column=7, value=row_data.get("category", ""))
			ws.cell(row=row_idx, column=8, value=row_data.get("quantity", 0))
			ws.cell(row=row_idx, column=9, value=format_datetime(row_data.get("sub_date"), "yyyy-MM-dd HH:mm") if row_data.get("sub_date") else "")
			ws.cell(row=row_idx, column=10, value=row_data.get("approval_status", ""))
			ws.cell(row=row_idx, column=11, value=row_data.get("submission_status", ""))
			ws.cell(row=row_idx, column=12, value=row_data.get("user", ""))
			ws.cell(row=row_idx, column=13, value=row_data.get("task_type", ""))
			ws.cell(row=row_idx, column=14, value=format_datetime(row_data.get("start_date"), "yyyy-MM-dd") if row_data.get("start_date") else "")
			ws.cell(row=row_idx, column=15, value=format_datetime(row_data.get("end_date"), "yyyy-MM-dd") if row_data.get("end_date") else "")

		# 调整列宽
		for col in ws.columns:
			max_length = 0
			column = col[0].column_letter
			for cell in col:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			ws.column_dimensions[column].width = adjusted_width

		# 保存文件
		file_name = f"数据查看_{frappe.utils.now()}.xlsx"
		file_path = os.path.join(get_site_path(), "public", "files", file_name)

		wb.save(file_path)

		# 返回文件 URL
		file_url = f"/files/{file_name}"

		return {
			"status": "success",
			"file_url": file_url,
			"message": "导出成功"
		}

	except Exception as e:
		frappe.log_error(title="导出数据失败", message=str(e))
		return {
			"status": "error",
			"message": str(e)
		}
