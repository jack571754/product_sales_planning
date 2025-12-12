"""
Import/Export API
导入导出相关API接口
从 store_detail.py 迁移
"""

import frappe
from product_sales_planning.utils.response_utils import success_response, error_response
from product_sales_planning.utils.validation_utils import validate_required_params


@frappe.whitelist()
def download_import_template(task_id=None):
	"""生成并下载Excel导入模板"""
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from frappe.utils.file_manager import save_file
		import io
		from datetime import datetime
		from product_sales_planning.utils.date_utils import get_next_n_months

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "商品计划导入模板"

		# 定义样式
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_font = Font(bold=True, color="FFFFFF", size=11)
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 生成月份列
		months = get_next_n_months(n=4, include_current=False)

		# 设置表头
		headers = ['产品编码', '产品名称'] + months
		ws.append(headers)

		# 设置表头样式
		for col_num in range(1, len(headers) + 1):
			cell = ws.cell(row=1, column=col_num)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.border = border

		# 添加示例数据
		example_row1 = ['PROD001', '示例商品A'] + [100 + i * 10 for i in range(len(months))]
		example_row2 = ['PROD002', '示例商品B'] + [50 + i * 5 for i in range(len(months))]

		for row_data in [example_row1, example_row2]:
			ws.append(row_data)
			for col_num in range(1, len(headers) + 1):
				cell = ws.cell(row=ws.max_row, column=col_num)
				cell.border = border
				if col_num > 2:
					cell.alignment = Alignment(horizontal='right', vertical='center')

		# 设置列宽
		ws.column_dimensions['A'].width = 15
		ws.column_dimensions['B'].width = 25
		for i in range(len(months)):
			col_letter = openpyxl.utils.get_column_letter(i + 3)
			ws.column_dimensions[col_letter].width = 12

		# 添加说明工作表
		ws_info = wb.create_sheet("填写说明")
		instructions = [
			["商品计划导入模板使用说明", ""],
			["", ""],
			["1. 基本要求", ""],
			["", "• 产品编码必须在系统中存在"],
			["", "• 产品名称仅用于参考，不影响导入"],
			["", "• 月份格式支持：2025-01、202501、2025/01"],
			["", "• 数量必须为整数，空值或0将被跳过"],
			["", ""],
			["2. 导入规则", ""],
			["", "• 如果记录已存在（相同店铺+任务+产品+月份），将更新数量"],
			["", "• 如果记录不存在，将创建新记录"],
			["", "• 导入前请确保已选择店铺和计划任务"],
			["", ""],
			["3. 注意事项", ""],
			["", "• 请勿修改表头行"],
			["", "• 建议单次导入不超过1000行数据"],
			["", "• 导入完成后请检查导入结果"],
		]

		for row_data in instructions:
			ws_info.append(row_data)

		ws_info.column_dimensions['A'].width = 20
		ws_info.column_dimensions['B'].width = 50
		title_cell = ws_info['A1']
		title_cell.font = Font(bold=True, size=14, color="4472C4")

		# 保存到内存
		file_content = io.BytesIO()
		wb.save(file_content)
		file_content.seek(0)

		filename = f"commodity_plan_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

		file_doc = save_file(
			fname=filename,
			content=file_content.read(),
			dt=None,
			dn=None,
			is_private=0
		)

		return success_response(
			file_url=file_doc.file_url,
			file_name=filename
		)

	except Exception as e:
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="生成导入模板失败", message=error_msg)
		return error_response(message=f"生成模板失败: {str(e)}")


@frappe.whitelist()
def import_commodity_data(store_id, task_id, file_url):
	"""从Excel导入商品计划数据"""
	try:
		import openpyxl
		from frappe.utils.file_manager import get_file_path
		from product_sales_planning.utils.date_utils import parse_month_string, get_month_first_day

		validate_required_params(
			{"store_id": store_id, "task_id": task_id},
			["store_id", "task_id"]
		)

		# 获取文件路径
		try:
			file_path = get_file_path(file_url)
		except Exception as e:
			return error_response(message=f"无法获取文件: {str(e)}")

		# 读取Excel文件
		try:
			wb = openpyxl.load_workbook(file_path, data_only=True)
			ws = wb.active
		except Exception as e:
			return error_response(message=f"无法读取Excel文件: {str(e)}")

		# 读取表头
		headers = []
		for cell in ws[1]:
			if cell.value:
				headers.append(str(cell.value).strip())
			else:
				headers.append("")

		if len(headers) < 3:
			return error_response(message="Excel格式错误：至少需要3列（产品编码、产品名称、月份数据）")

		# 解析月份列
		month_columns = [h for h in headers[2:] if h]

		if not month_columns:
			return error_response(message="Excel格式错误：未找到月份列")

		inserted_count = 0
		updated_count = 0
		errors = []
		skipped_count = 0

		# 从第2行开始读取数据
		for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			if not row or not row[0]:
				skipped_count += 1
				continue

			try:
				product_code = str(row[0]).strip()
			except:
				errors.append(f"第{row_idx}行: 产品编码格式错误")
				continue

			# 验证产品是否存在
			if not frappe.db.exists("Product List", product_code):
				errors.append(f"第{row_idx}行: 产品编码 {product_code} 不存在")
				continue

			# 处理每个月份的数据
			for col_idx, month_str in enumerate(month_columns):
				try:
					if len(row) <= 2 + col_idx:
						continue

					quantity_value = row[2 + col_idx]

					if quantity_value is None or quantity_value == '' or quantity_value == 0:
						continue

					try:
						quantity = int(float(quantity_value))
					except:
						errors.append(f"第{row_idx}行-{month_str}: 数量格式错误 ({quantity_value})")
						continue

					# 解析月份
					parsed_month = parse_month_string(month_str)
					if not parsed_month:
						errors.append(f"第{row_idx}行: 月份格式错误 ({month_str})")
						continue

					sub_date = get_month_first_day(parsed_month)

					# 检查记录是否存在
					filters = {
						"store_id": store_id,
						"task_id": task_id,
						"code": product_code,
						"sub_date": sub_date
					}

					existing = frappe.db.get_value("Commodity Schedule", filters, "name")

					if existing:
						frappe.db.set_value("Commodity Schedule", existing, "quantity", quantity)
						updated_count += 1
					else:
						new_doc = frappe.new_doc("Commodity Schedule")
						new_doc.store_id = store_id
						new_doc.task_id = task_id
						new_doc.code = product_code
						new_doc.quantity = quantity
						new_doc.sub_date = sub_date
						new_doc.insert()
						inserted_count += 1

				except Exception as inner_e:
					errors.append(f"第{row_idx}行-{month_str}: {str(inner_e)}")

		frappe.db.commit()

		msg = f"成功导入 {inserted_count} 条，更新 {updated_count} 条"
		if skipped_count > 0:
			msg += f"，跳过 {skipped_count} 行空数据"

		return success_response(
			message=msg,
			inserted=inserted_count,
			updated=updated_count,
			skipped=skipped_count,
			errors=errors[:20]
		)

	except Exception as e:
		frappe.db.rollback()
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="Excel导入失败", message=error_msg)
		return error_response(message=f"导入失败: {str(e)}")


@frappe.whitelist()
def export_commodity_data(store_id=None, task_id=None):
	"""导出商品计划数据到Excel"""
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from frappe.utils.file_manager import save_file
		import io
		from datetime import datetime
		from product_sales_planning.services.commodity_service import CommodityScheduleService

		# 获取数据
		result = CommodityScheduleService.get_commodity_data(
			store_id=store_id,
			task_id=task_id,
			start=0,
			page_length=10000,
			view_mode="multi"
		)

		if not result.get("data"):
			return error_response(message="没有数据可导出")

		data = result["data"]
		months = result.get("months", [])

		# 创建工作簿
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "商品计划数据"

		# 定义样式
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_font = Font(bold=True, color="FFFFFF", size=11)
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 设置表头
		headers = ['产品编码', '产品名称', '规格', '品牌', '类别'] + months
		ws.append(headers)

		# 设置表头样式
		for col_num in range(1, len(headers) + 1):
			cell = ws.cell(row=1, column=col_num)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.border = border

		# 写入数据
		for item in data:
			row_data = [
				item.get('code', ''),
				item.get('name1', ''),
				item.get('specifications', ''),
				item.get('brand', ''),
				item.get('category', '')
			]

			# 添加月份数据
			for month in months:
				month_data = item.get('months', {}).get(month, {})
				quantity = month_data.get('quantity', 0) if month_data else 0
				row_data.append(quantity)

			ws.append(row_data)

			# 设置数据行样式
			for col_num in range(1, len(headers) + 1):
				cell = ws.cell(row=ws.max_row, column=col_num)
				cell.border = border
				if col_num > 5:
					cell.alignment = Alignment(horizontal='right', vertical='center')

		# 设置列宽
		ws.column_dimensions['A'].width = 15
		ws.column_dimensions['B'].width = 25
		ws.column_dimensions['C'].width = 12
		ws.column_dimensions['D'].width = 12
		ws.column_dimensions['E'].width = 12
		for i in range(len(months)):
			col_letter = openpyxl.utils.get_column_letter(i + 6)
			ws.column_dimensions[col_letter].width = 12

		# 保存到内存
		file_content = io.BytesIO()
		wb.save(file_content)
		file_content.seek(0)

		# 生成文件名
		filename_parts = ["commodity_plan_export"]
		if store_id:
			filename_parts.append(f"store_{store_id}")
		if task_id:
			filename_parts.append(f"task_{task_id}")
		filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
		filename = "_".join(filename_parts) + ".xlsx"

		# 保存文件
		file_doc = save_file(
			fname=filename,
			content=file_content.read(),
			dt=None,
			dn=None,
			is_private=0
		)

		return success_response(
			file_url=file_doc.file_url,
			file_name=filename,
			record_count=len(data)
		)

	except Exception as e:
		import traceback
		error_msg = traceback.format_exc()
		frappe.log_error(title="导出Excel失败", message=error_msg)
		return error_response(message=f"导出失败: {str(e)}")


@frappe.whitelist()
def download_mechanism_template():
	"""下载机制导入模板（占位函数）"""
	return error_response(message="功能开发中")


@frappe.whitelist()
def import_mechanism_excel(file_url):
	"""导入机制Excel（占位函数）"""
	return error_response(message="功能开发中")